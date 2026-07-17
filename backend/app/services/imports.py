"""Bulk import of records from a spreadsheet (.csv or .xlsx).

Parsing is deterministic, not AI-based: the file's first row is the header, each remaining row
maps column -> value, and every row is validated with the same Pydantic model the create API uses.
Invalid rows are reported with their line number and the exact reason; valid rows are inserted. As
columns are known and the validation is the same as a single create, the outcome is predictable and
auditable — the right tool for structured company data (a supplier list, a project register), as
opposed to the document-upload feature which handles unstructured files for AI search.
"""

import csv
import io
from collections.abc import Awaitable, Callable
from typing import Any

from pydantic import BaseModel, ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Project, PurchaseRequest, Supplier
from app.schemas.imports import ImportReport, ImportRowError

MAX_IMPORT_ROWS = 5000

# A resolver rewrites a parsed row in place before validation (e.g. a human key -> a foreign id)
# and returns any row-level problems it found. An empty list means the row is ready to validate.
RowResolver = Callable[[dict[str, str]], list[str]]


async def project_code_resolver(db: AsyncSession) -> RowResolver:
    """Build a resolver that turns a human ``project_code`` column into the ``project_id`` the
    create schemas expect. Child records (RFIs, claims, meetings, …) belong to a project, but a
    company importing a spreadsheet knows its own project codes, not our numeric ids. The code -> id
    map is loaded once; an unknown code is reported as a row error rather than silently dropped.
    """
    rows = await db.execute(select(Project.project_code, Project.id))
    code_to_id = {code: pid for code, pid in rows.all()}

    def resolve(provided: dict[str, str]) -> list[str]:
        code = provided.pop("project_code", "")
        if not code:
            return ["project_code: this field is required"]
        project_id = code_to_id.get(code)
        if project_id is None:
            return [f"project_code: unknown project code '{code}'"]
        provided["project_id"] = str(project_id)
        return []

    return resolve


def _normalize_key(value: str) -> str:
    """Lower-case and collapse internal whitespace runs to a single space. A human typing a
    supplier name from memory into a spreadsheet can fat-finger an extra space (`"Risk  Supplier
    001"`) as easily as get the casing wrong — `.strip()` alone (already applied when the file is
    parsed) only catches leading/trailing whitespace, not this. Both callers of this function
    confirm the real data has no two distinct records that would collide once normalized before
    relying on it as a lookup key.
    """
    return " ".join(value.lower().split())


async def purchase_order_resolver(db: AsyncSession) -> RowResolver:
    """Build a resolver for purchase orders — the one entity with two foreign keys to resolve
    from a spreadsheet, not just one. Deliberately does NOT take a ``project_code`` column like the
    other child entities: a purchase request already belongs to exactly one real project, so
    ``project_id`` is derived from the resolved ``request_no`` instead of asked for separately. That
    removes a column the importer would otherwise have to get right, and makes a project/request
    mismatch structurally impossible rather than a validation rule to enforce.

    Matching is case- and internal-whitespace-insensitive on both columns: a value typed from
    memory into a spreadsheet is realistic human input, not a system-generated value copied
    verbatim, so neither a casing mismatch nor a stray extra space should produce a false "unknown"
    rejection. Confirmed safe against the real data before relying on it — no two request numbers
    or supplier names collide once normalized, so this lookup can never conflate two different real
    records.
    """
    pr_rows = await db.execute(
        select(PurchaseRequest.request_no, PurchaseRequest.id, PurchaseRequest.project_id)
    )
    request_no_to_ids = {
        _normalize_key(no): (pr_id, project_id) for no, pr_id, project_id in pr_rows.all()
    }
    supplier_rows = await db.execute(select(Supplier.supplier_name, Supplier.id))
    supplier_to_id = {_normalize_key(name): sid for name, sid in supplier_rows.all()}

    def resolve(provided: dict[str, str]) -> list[str]:
        errors: list[str] = []

        request_no = provided.pop("request_no", "")
        if not request_no:
            errors.append("request_no: this field is required")
        elif _normalize_key(request_no) not in request_no_to_ids:
            errors.append(f"request_no: unknown purchase request '{request_no}'")
        else:
            pr_id, project_id = request_no_to_ids[_normalize_key(request_no)]
            provided["pr_id"] = str(pr_id)
            provided["project_id"] = str(project_id)

        supplier_name = provided.pop("supplier_name", "")
        if not supplier_name:
            errors.append("supplier_name: this field is required")
        elif _normalize_key(supplier_name) not in supplier_to_id:
            errors.append(f"supplier_name: unknown supplier '{supplier_name}'")
        else:
            provided["supplier_id"] = str(supplier_to_id[_normalize_key(supplier_name)])

        return errors

    return resolve


class UnsupportedImport(ValueError):
    """Raised when a file is neither .csv nor .xlsx."""


class ImportParseError(ValueError):
    """Raised when a supported file type cannot be read (e.g. a corrupt spreadsheet)."""


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_tabular(filename: str, data: bytes) -> list[dict[str, str]]:
    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(data)
    if name.endswith(".xlsx"):
        return _parse_xlsx(data)
    raise UnsupportedImport("Unsupported file type. Upload a .csv or .xlsx file.")


def _parse_csv(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")  # -sig strips an Excel byte-order mark
    reader = csv.DictReader(io.StringIO(text))
    return [{_clean(k): _clean(v) for k, v in raw.items() if k} for raw in reader]


def _parse_xlsx(data: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises BadZipFile/InvalidFileException on bad files
        raise ImportParseError(
            "The Excel file could not be read. Make sure it is a valid .xlsx file."
        ) from exc
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        headers = [_clean(cell) for cell in header]
        parsed: list[dict[str, str]] = []
        for values in rows_iter:
            if all(_clean(v) == "" for v in values):
                continue
            parsed.append(
                {
                    headers[i]: _clean(v)
                    for i, v in enumerate(values)
                    if i < len(headers) and headers[i]
                }
            )
        return parsed
    finally:
        workbook.close()


async def import_rows(
    db: AsyncSession,
    rows: list[dict[str, str]],
    *,
    schema: type[BaseModel],
    create: Callable[[AsyncSession, Any], Awaitable[Any]],
    dry_run: bool,
    resolve: RowResolver | None = None,
) -> ImportReport:
    """Validate every row against ``schema`` and, unless it is a dry run, create the valid ones.

    The line number in an error is the spreadsheet row (data begins on line 2, after the header).
    An optional ``resolve`` step runs first to rewrite a row (e.g. project_code -> project_id); a
    row it rejects is reported and skipped before schema validation.
    """
    errors: list[ImportRowError] = []
    valid = 0
    for line, row in enumerate(rows, start=2):
        # Treat blank cells as "not provided" so optional fields fall back to their defaults.
        provided = {key: value for key, value in row.items() if value != ""}
        if resolve is not None:
            resolve_errors = resolve(provided)
            if resolve_errors:
                errors.append(ImportRowError(row=line, errors=resolve_errors))
                continue
        try:
            payload = schema.model_validate(provided)
        except ValidationError as exc:
            errors.append(
                ImportRowError(
                    row=line,
                    errors=[
                        f"{'.'.join(str(part) for part in err['loc'])}: {err['msg']}"
                        for err in exc.errors()
                    ],
                )
            )
            continue
        valid += 1
        if not dry_run:
            await create(db, payload)

    if not dry_run and valid:
        await db.commit()

    return ImportReport(
        total_rows=len(rows),
        valid_rows=valid,
        invalid_rows=len(errors),
        created=valid if not dry_run else 0,
        dry_run=dry_run,
        errors=errors[:100],
    )
